import openpyxl

#Criar uma planila (book)
book = openpyxl.Workbook()

#Visualizar paginas existentes
print(book.sheetnames)

#como criar um pagina
book.create_sheet('Frutas')

#Selecionar pagina
frutas_page = book['Frutas']



#add dados
frutas_page.append(['Frutas','qtd','preço'])
frutas_page.append(['bananas','5','3,90'])
frutas_page.append(['bananas2','45','2,50'])
frutas_page.append(['bananas3','8','33,00'])
frutas_page.append(['bananas4','2','55,50'])

#Slavar planilha
book.save('Planilha de compras.xlsx')
